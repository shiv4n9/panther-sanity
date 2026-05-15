#!/usr/bin/env python3
"""
Panther Sanity — API Server
- Serves CSV files from the mounted CSV directory
- Auto-watches for new CSVs and ingests into PostgreSQL
- Exposes manual ingest trigger endpoint
- Provides 30-day historical data endpoint for the frontend
"""

from flask import Flask, send_file, jsonify, request
from flask_cors import CORS
import os
import re
import time
import threading
from pathlib import Path
from datetime import datetime, date, timedelta
import logging
import psycopg2
import psycopg2.extras
import psycopg2.pool
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ─── Logging ──────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ─── App Config ───────────────────────────────────────────────
app = Flask(__name__)
CORS(app)

CSV_DIR      = os.environ.get('CSV_DIR', '/data/csvs')
PORT         = int(os.environ.get('PORT', 3001))
HOST         = os.environ.get('HOST', '0.0.0.0')
DATABASE_URL = os.environ.get('DATABASE_URL', '')

os.makedirs(CSV_DIR, exist_ok=True)

# ─── Database Helpers ─────────────────────────────────────────
_db_pool = None


def init_db_pool(minconn=2, maxconn=10):
    """Initialize the connection pool. Call once after DB is reachable."""
    global _db_pool
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL environment variable not set")
    _db_pool = psycopg2.pool.ThreadedConnectionPool(minconn, maxconn, DATABASE_URL)
    logger.info(f"Database connection pool initialized (min={minconn}, max={maxconn})")


def get_db():
    """Get a connection from the pool (or create a one-off if pool not ready)."""
    if _db_pool:
        conn = _db_pool.getconn()
        conn.autocommit = False
        return conn
    # Fallback for startup / health checks before pool is ready
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL environment variable not set")
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    return conn


def put_db(conn):
    """Return a connection to the pool (or close it if no pool)."""
    if _db_pool:
        _db_pool.putconn(conn)
    else:
        conn.close()


def wait_for_db(retries=10, delay=3):
    """Block until Postgres is reachable (gives the db container time to start)."""
    for i in range(retries):
        try:
            conn = get_db()
            conn.close()
            logger.info("Database connection established.")
            return True
        except Exception as e:
            logger.warning(f"Waiting for database... ({i + 1}/{retries}): {e}")
            time.sleep(delay)
    logger.error("Could not connect to database after multiple retries.")
    return False


# ─── CSV Parsing ──────────────────────────────────────────────
def parse_csv_line(line: str) -> list:
    """Parse a CSV line, respecting quoted fields."""
    fields, current, in_quotes = [], '', False
    for char in line:
        if char == '"':
            in_quotes = not in_quotes
        elif char == ',' and not in_quotes:
            fields.append(current.strip())
            current = ''
        else:
            current += char
    fields.append(current.strip())
    return fields


def parse_csv_file(file_path: str) -> dict:
    """Parse a sanity CSV file into metadata + rows."""
    with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
        lines = [l.rstrip('\n') for l in f.readlines()]

    if len(lines) < 3:
        raise ValueError(f"CSV too short: {file_path}")

    platform   = lines[0].split(',')[1].strip() if ',' in lines[0] else 'Unknown'
    image_name = lines[1].split(',')[1].strip() if ',' in lines[1] else 'Unknown'

    rows = []
    for line in lines[3:]:
        line = line.strip()
        if not line:
            continue
        fields = parse_csv_line(line)
        if len(fields) < 3:
            continue

        # Test case names may contain unquoted commas (e.g. "IPSEC VPN Throughput (S2S, PSK, AES256-GCM)")
        # Always use LAST field as throughput, SECOND-TO-LAST as parameter,
        # and join everything before as the test case name.
        throughput = fields[-1].strip()
        parameter  = fields[-2].strip()
        test_case  = ','.join(f.strip() for f in fields[:-2]).strip()

        # Skip the column header row
        if test_case.upper() == 'TESTCASE':
            continue
        if not test_case:
            continue

        cpu_match = re.search(r'CPU:\s*(\d+)%', throughput, re.IGNORECASE)
        cpu = (cpu_match.group(1) + '%') if cpu_match else None

        rows.append({
            'test_case':  test_case,
            'parameter':  parameter,
            'throughput': throughput,
            'cpu':        cpu,
            'memory':     None,   # extend when source data has it
            'shm':        None,
        })

    return {
        'platform':   platform,
        'image_name': image_name,
        'rows':       rows,
    }


def ingest_csv(file_path: str, force: bool = False) -> dict:
    """
    Parse a CSV file and insert its rows into sanity_runs.
    Skips rows that already exist for this csv_filename + run_date.
    Returns a summary dict.

    Args:
        file_path: Path to the CSV file.
        force: If True, delete existing rows and re-insert.
    
    Note: The date in the filename represents the image build date, not the test execution date.
    We use the file modification time as the actual test run date.
    """
    path     = Path(file_path)
    filename = path.name

    # Use file modification time as the run date (when test was actually executed)
    # The date in filename is the image build date, not test execution date
    mtime    = path.stat().st_mtime
    run_date = datetime.fromtimestamp(mtime).date()

    logger.info(f"Ingesting {filename} (run_date={run_date}, force={force})")

    parsed  = parse_csv_file(file_path)
    rows    = parsed['rows']
    inserted = 0
    skipped  = 0

    conn = get_db()
    try:
        with conn.cursor() as cur:
            # Check if this file has already been ingested for this date
            cur.execute(
                "SELECT COUNT(*) FROM sanity_runs WHERE csv_filename=%s AND run_date=%s",
                (filename, run_date)
            )
            existing = cur.fetchone()[0]
            if existing > 0 and not force:
                logger.info(f"Already ingested {filename} for {run_date}, skipping.")
                return {'status': 'skipped', 'filename': filename, 'reason': 'already ingested'}

            if force and existing > 0:
                 # Delete existing rows for this file so they get cleanly re-inserted
                 cur.execute("DELETE FROM sanity_runs WHERE csv_filename=%s AND run_date=%s", (filename, run_date))

            for row in rows:
                # Skip rows where throughput is a GNATS PR number (6+ digit
                # pure numeric value). Short numbers like 744, 2100 are
                # legitimate throughput values.
                throughput_val = row['throughput'].strip()
                if throughput_val.isdigit() and len(throughput_val) >= 6:
                    logger.debug(f"Skipping PR row: {row['test_case']} → {throughput_val}")
                    skipped += 1
                    continue

                cur.execute(
                    """
                    INSERT INTO sanity_runs
                      (run_date, test_case, parameter, throughput, cpu, memory, shm,
                       platform, image_name, csv_filename)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        run_date,
                        row['test_case'],
                        row['parameter'],
                        row['throughput'],
                        row['cpu'],
                        row['memory'],
                        row['shm'],
                        parsed['platform'],
                        parsed['image_name'],
                        filename,
                    )
                )
                inserted += 1

        conn.commit()
        logger.info(f"Ingested {inserted} rows from {filename}")
        return {
            'status':   'success',
            'filename': filename,
            'run_date': str(run_date),
            'inserted': inserted,
            'skipped':  skipped,
        }
    except Exception as e:
        conn.rollback()
        logger.error(f"Ingest failed for {filename}: {e}")
        raise
    finally:
        put_db(conn)


# ─── Watchdog ─────────────────────────────────────────────────
class CSVHandler(FileSystemEventHandler):
    """Auto-ingest newly created or moved CSV files."""

    def _handle(self, event):
        if event.is_directory:
            return
        src = getattr(event, 'dest_path', event.src_path)
        if src.endswith('.csv'):
            logger.info(f"Watchdog detected: {src}")
            try:
                ingest_csv(src, force=False)
            except Exception as e:
                logger.error(f"Auto-ingest error: {e}")

    on_created = _handle
    on_moved   = _handle


def start_watcher():
    handler  = CSVHandler()
    observer = Observer()
    observer.schedule(handler, CSV_DIR, recursive=False)
    observer.start()
    logger.info(f"Watchdog watching: {CSV_DIR}")
    return observer


# ─── Routes ───────────────────────────────────────────────────

@app.route('/health', methods=['GET'])
def health_check():
    db_ok = False
    try:
        conn  = get_db()
        put_db(conn)
        db_ok = True
    except Exception:
        pass

    return jsonify({
        'status':       'healthy' if db_ok else 'degraded',
        'db_connected': db_ok,
        'csv_dir':      CSV_DIR,
        'timestamp':    datetime.now().isoformat()
    })


@app.route('/api/sanity-results/latest', methods=['GET'])
def get_latest():
    """Serve the most recent CSV file (raw download)."""
    try:
        csv_files = sorted(
            Path(CSV_DIR).glob('*.csv'),
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )
        if not csv_files:
            return jsonify({'error': 'No CSV files found'}), 404

        latest = csv_files[0]
        logger.info(f"Serving latest: {latest.name}")
        return send_file(latest, mimetype='text/csv', as_attachment=False,
                         download_name=latest.name)
    except Exception as e:
        logger.error(f"Error serving latest: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/sanity-results', methods=['GET'])
def list_files():
    """List all CSV files with metadata."""
    try:
        files = []
        for p in sorted(Path(CSV_DIR).glob('*.csv'),
                        key=lambda x: x.stat().st_mtime, reverse=True):
            st = p.stat()
            files.append({
                'name':     p.name,
                'size':     st.st_size,
                'modified': datetime.fromtimestamp(st.st_mtime).isoformat(),
            })
        return jsonify({'files': files, 'count': len(files)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sanity-results/<filename>', methods=['GET'])
def get_file(filename):
    """Serve a specific CSV file."""
    try:
        filename  = os.path.basename(filename)
        file_path = os.path.join(CSV_DIR, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': f'Not found: {filename}'}), 404
        if not filename.endswith('.csv'):
            return jsonify({'error': 'Only CSV files allowed'}), 400
        return send_file(file_path, mimetype='text/csv', as_attachment=False,
                         download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ingest', methods=['POST'])
def manual_ingest():
    """
    Manual ingest trigger.
    POST with no body → ingests the latest CSV.
    POST with JSON {"filename": "foo.csv"} → ingests that specific file.
    """
    try:
        body     = request.get_json(silent=True) or {}
        filename = body.get('filename')

        if filename:
            file_path = os.path.join(CSV_DIR, os.path.basename(filename))
        else:
            # Pick the latest CSV
            csv_files = sorted(
                Path(CSV_DIR).glob('*.csv'),
                key=lambda x: x.stat().st_mtime,
                reverse=True
            )
            if not csv_files:
                return jsonify({'error': 'No CSV files found'}), 404
            file_path = str(csv_files[0])

        if not os.path.exists(file_path):
            return jsonify({'error': f'File not found: {file_path}'}), 404

        force = request.args.get('force', 'false').lower() == 'true'
        result = ingest_csv(file_path, force=force)
        return jsonify(result), 200

    except Exception as e:
        logger.error(f"Manual ingest error: {e}")
        return jsonify({'error': str(e)}), 500


# ─── XLSX Parsing Helpers ─────────────────────────────────────

def _is_section_header_row(col_b):
    """Detect section-header rows (same logic as frontend xlsxParser.js)."""
    if not col_b:
        return False
    clean = col_b.replace('\r', ' ').replace('\n', ' ').strip()
    if clean and clean[0].isdigit():
        return False
    lower = clean.lower()
    return any(kw in lower for kw in ['throughput', 'cps', 'scale', 'tps'])


def _extract_release(cell_value):
    """Extract release version from Row 0, Col A."""
    if not cell_value:
        return 'Unknown'
    m = re.search(r'Release:\s*(.+?)(?:\r?\n|$)', str(cell_value))
    return m.group(1).strip() if m else 'Unknown'


def _parse_xlsx_sheet(ws):
    """Parse a single openpyxl worksheet into sections (mirrors JS parseSheet)."""
    # Row 1 = header, rows 2+ = data (openpyxl is 1-indexed)
    release = _extract_release(ws.cell(row=1, column=1).value)

    sections = []
    current_section = None
    col_mapping = None

    for row in ws.iter_rows(min_row=2, max_col=10, values_only=False):
        cells = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
        col_a, col_b = cells[0], cells[1]

        if not col_a and not col_b:
            continue

        if _is_section_header_row(col_b):
            current_section = {'category': col_a, 'tests': []}
            sections.append(current_section)
            # Determine column semantics
            col_d = cells[3].lower() if len(cells) > 3 else ''
            d_is_comments = 'comment' in col_d or 'session' in col_d
            col_mapping = {
                'throughput': 1, 'cpu': 2,
                'shm': -1 if d_is_comments else 3,
                'session': 3 if d_is_comments else -1,
            }
            continue

        if not current_section:
            current_section = {'category': 'General', 'tests': []}
            sections.append(current_section)
            col_mapping = {'throughput': 1, 'cpu': 2, 'shm': 3, 'session': -1}

        throughput = cells[col_mapping['throughput']] if col_mapping['throughput'] >= 0 else ''
        cpu = cells[col_mapping['cpu']] if col_mapping['cpu'] >= 0 else ''
        shm = cells[col_mapping['shm']] if col_mapping['shm'] >= 0 else ''

        # Comments
        comments = ''
        if col_mapping['session'] >= 0:
            comments = cells[col_mapping['session']]
            extra = cells[4] if len(cells) > 4 else ''
            if extra:
                comments = f"{comments} — {extra}" if comments else extra
        else:
            comments = cells[4] if len(cells) > 4 else ''

        current_section['tests'].append({
            'test_case': col_a,
            'throughput': throughput,
            'cpu': f"{cpu}%" if cpu and not cpu.endswith('%') else cpu,
            'shm': f"{shm}%" if shm and not shm.endswith('%') else shm,
            'comments': comments,
        })

    return {'release': release, 'sections': sections}


def _get_previous_snapshot(conn, platform):
    """Get the most recent throughput values for a platform to compute diffs."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT DISTINCT ON (test_case, category)
                test_case, category, throughput, run_date
            FROM sanity_runs
            WHERE platform = %s
            ORDER BY test_case, category, run_date DESC
        """, (platform,))
        return {f"{r['category']}::{r['test_case']}": r for r in cur.fetchall()}


@app.route('/api/ingest-xlsx', methods=['POST'])
def ingest_xlsx():
    """
    Parse the SRX4XX_Datasheet.xlsx and insert all rows into the database.
    Also computes a changelog diff against the previous snapshot.
    """
    try:
        import openpyxl

        body = request.get_json(silent=True) or {}
        xlsx_path = body.get('path', '/data/xlsx/SRX4XX_Datasheet.xlsx')

        # Also check common local dev paths
        if not os.path.exists(xlsx_path):
            for alt in [
                os.path.join(os.path.dirname(__file__), '..', 'public', 'data', 'SRX4XX_Datasheet.xlsx'),
                '/app/public/data/SRX4XX_Datasheet.xlsx',
            ]:
                if os.path.exists(alt):
                    xlsx_path = alt
                    break

        if not os.path.exists(xlsx_path):
            return jsonify({'error': f'XLSX not found: {xlsx_path}'}), 404

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        run_date = date.today()
        filename = os.path.basename(xlsx_path)

        total_inserted = 0
        total_updated = 0
        total_added = 0
        diff_data = {'updated': [], 'added': []}
        releases = {}

        conn = get_db()
        try:
            # First pass: parse all sheets to collect releases for the log entry
            parsed_sheets = {}
            for sheet_name in ['SRX400', 'SRX440']:
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found, skipping")
                    continue
                ws = wb[sheet_name]
                parsed_sheets[sheet_name] = _parse_xlsx_sheet(ws)
                releases[sheet_name.lower()] = parsed_sheets[sheet_name]['release']

            # Create ingestion_log entry FIRST to get a unique ID
            import json as json_mod
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO ingestion_log
                      (release_400, release_440, source_file, tests_added, tests_updated, diff_json)
                    VALUES (%s, %s, %s, 0, 0, '{}'::jsonb)
                    RETURNING id
                """, (
                    releases.get('srx400', ''),
                    releases.get('srx440', ''),
                    filename,
                ))
                ingestion_id = cur.fetchone()[0]

            # Second pass: compute diffs and insert rows
            for sheet_name, parsed in parsed_sheets.items():
                platform = sheet_name

                # Get previous snapshot for diff (excludes current ingestion)
                prev = _get_previous_snapshot(conn, platform)

                with conn.cursor() as cur:
                    for section in parsed['sections']:
                        category = section['category']
                        for test in section['tests']:
                            if not test['test_case'] or not test['throughput']:
                                continue

                            key = f"{category}::{test['test_case']}"
                            prev_row = prev.get(key)

                            if prev_row:
                                if prev_row['throughput'] != test['throughput']:
                                    diff_data['updated'].append({
                                        'test_case': test['test_case'],
                                        'category': category,
                                        'platform': platform,
                                        'old_value': prev_row['throughput'],
                                        'new_value': test['throughput'],
                                    })
                                    total_updated += 1
                            else:
                                diff_data['added'].append({
                                    'test_case': test['test_case'],
                                    'category': category,
                                    'platform': platform,
                                    'value': test['throughput'],
                                })
                                total_added += 1

                            cur.execute("""
                                INSERT INTO sanity_runs
                                  (run_date, test_case, parameter, throughput, cpu, memory, shm,
                                   platform, image_name, csv_filename, release, category)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """, (
                                run_date,
                                test['test_case'],
                                category,
                                test['throughput'],
                                test['cpu'],
                                None,
                                test['shm'],
                                platform,
                                parsed['release'],
                                filename,
                                parsed['release'],
                                category,
                            ))
                            total_inserted += 1

            # Update the ingestion_log with actual diff counts
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE ingestion_log
                    SET tests_added = %s, tests_updated = %s, diff_json = %s
                    WHERE id = %s
                """, (total_added, total_updated, json_mod.dumps(diff_data), ingestion_id))

            conn.commit()
            wb.close()

            return jsonify({
                'status': 'success',
                'run_date': str(run_date),
                'ingestion_id': ingestion_id,
                'inserted': total_inserted,
                'updated': total_updated,
                'added': total_added,
                'releases': releases,
                'diff': diff_data,
            }), 200

        except Exception as e:
            conn.rollback()
            raise
        finally:
            put_db(conn)

    except Exception as e:
        logger.error(f"XLSX ingest error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/changelog', methods=['GET'])
def get_changelog():
    """Return recent ingestion logs with diffs."""
    limit = int(request.args.get('limit', 5))
    try:
        conn = get_db()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT id, ingested_at, release_400, release_440, source_file,
                           tests_added, tests_updated, diff_json
                    FROM ingestion_log
                    ORDER BY ingested_at DESC
                    LIMIT %s
                """, (limit,))
                rows = cur.fetchall()
        finally:
            put_db(conn)

        result = []
        for row in rows:
            result.append({
                'id': row['id'],
                'ingested_at': row['ingested_at'].isoformat() if row['ingested_at'] else None,
                'release_400': row['release_400'],
                'release_440': row['release_440'],
                'source_file': row['source_file'],
                'tests_added': row['tests_added'],
                'tests_updated': row['tests_updated'],
                'diff': row['diff_json'] if isinstance(row['diff_json'], dict) else {},
            })
        return jsonify({'changelog': result})

    except Exception as e:
        logger.error(f"Changelog query error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/sanity-history', methods=['GET'])
def get_sanity_history():
    """
    Historical data for a specific test case across all ingestions.
    Query: ?test_case=...&platform=...&category=...&days=90
    Returns time-series data for charting.
    """
    test_case = request.args.get('test_case', '')
    platform = request.args.get('platform', '')
    category = request.args.get('category', '')
    days = int(request.args.get('days', 90))

    if not test_case or not platform:
        return jsonify({'error': 'test_case and platform are required'}), 400

    since = date.today() - timedelta(days=days)

    try:
        conn = get_db()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                query = """
                    SELECT run_date, throughput, cpu, shm, release, image_name, created_at
                    FROM sanity_runs
                    WHERE test_case = %s AND platform = %s AND run_date >= %s
                """
                params = [test_case, platform, since]

                if category:
                    query += " AND category = %s"
                    params.append(category)

                query += " ORDER BY created_at ASC"
                cur.execute(query, params)
                rows = cur.fetchall()
        finally:
            put_db(conn)

        history = []
        for row in rows:
            history.append({
                'date': row['created_at'].strftime('%Y-%m-%d %H:%M') if row['created_at'] else str(row['run_date']),
                'throughput': row['throughput'],
                'throughput_numeric': _extract_throughput_value(row['throughput']),
                'cpu': row['cpu'] or '',
                'shm': row['shm'] or '',
                'release': row['release'] or '',
                'image_name': row['image_name'] or '',
            })

        return jsonify({
            'test_case': test_case,
            'platform': platform,
            'category': category,
            'days': days,
            'count': len(history),
            'history': history,
        })

    except Exception as e:
        logger.error(f"Sanity history error: {e}")
        return jsonify({'error': str(e)}), 500



@app.route('/api/history', methods=['GET'])
def get_history():
    """
    30-day historical data for a specific test case + parameter.
    Query params: ?test_case=...&parameter=...&days=30
    Returns daily aggregated throughput, cpu, memory, shm values.
    """
    test_case = request.args.get('test_case', '')
    parameter = request.args.get('parameter', '')
    days      = int(request.args.get('days', 30))

    if not test_case or not parameter:
        return jsonify({'error': 'test_case and parameter are required'}), 400

    since = date.today() - timedelta(days=days)

    try:
        conn = get_db()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                      run_date,
                      throughput,
                      cpu,
                      memory,
                      shm,
                      platform,
                      image_name
                    FROM sanity_runs
                    WHERE test_case = %s
                      AND parameter = %s
                      AND run_date  >= %s
                    ORDER BY run_date ASC
                    """,
                    (test_case, parameter, since)
                )
                rows = cur.fetchall()
        finally:
            put_db(conn)

        # Build response — one point per day
        history = []
        for i, row in enumerate(rows):
            history.append({
                'day':        f"Day {i + 1}",
                'date':       str(row['run_date']),
                'throughput': _extract_throughput_value(row['throughput']),
                'cpu':        row['cpu'] or 'N/A',
                'memory':     row['memory'] or 'N/A',
                'shm':        row['shm'] or 'N/A',
                'platform':   row['platform'],
                'image_name': row['image_name'],
            })

        return jsonify({
            'test_case': test_case,
            'parameter': parameter,
            'days':      days,
            'count':     len(history),
            'history':   history,
        })

    except Exception as e:
        logger.error(f"History query error: {e}")
        return jsonify({'error': str(e)}), 500


def _extract_throughput_value(throughput_str: str) -> float:
    """
    Extract a numeric Gbps value from a throughput string like
    '1.8 Gbps, CPU: 85%' or '1.75'.
    """
    if not throughput_str:
        return 0.0
    # Strip CPU annotation
    clean = re.sub(r',?\s*CPU:\s*\d+%', '', throughput_str, flags=re.IGNORECASE).strip()
    # Find first float
    m = re.search(r'[\d.]+', clean)
    return float(m.group()) if m else 0.0


@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint not found'}), 404


@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': 'Internal server error'}), 500


# ─── Entry Point ──────────────────────────────────────────────
if __name__ == '__main__':
    # Wait for Postgres before starting the server
    if DATABASE_URL:
        wait_for_db()
        init_db_pool()
    else:
        logger.warning("DATABASE_URL not set — running without database support")

    # Start watchdog in a background thread
    observer = start_watcher()

    logger.info(f"Starting Panther Sanity API on http://{HOST}:{PORT}")
    try:
        app.run(
            host=HOST,
            port=PORT,
            debug=os.environ.get('DEBUG', 'false').lower() == 'true',
            use_reloader=False   # disable reloader so watchdog thread survives
        )
    finally:
        observer.stop()
        observer.join()
